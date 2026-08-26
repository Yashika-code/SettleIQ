import pandas as pd
import numpy as np
import datetime
import os
from exception_explainer import explain_exception

class ReconciliationEngine:
    def __init__(self, razorpay_csv='razorpay_settlements.csv', bank_csv='bank_statement.csv', gst_csv='gst_records.csv'):
        self.razorpay_csv = razorpay_csv
        self.bank_csv = bank_csv
        self.gst_csv = gst_csv
        
        self.razorpay_df = pd.DataFrame()
        self.bank_df = pd.DataFrame()
        self.gst_df = pd.DataFrame()
        
        self.matched_pairs = []
        self.exceptions = []
        
    def load_data(self):
        """Load input CSVs into DataFrames with proper date types and column normalization."""
        if (not os.path.exists(self.razorpay_csv) or not os.path.exists(self.bank_csv) or 
            os.path.getsize(self.razorpay_csv) == 0 or os.path.getsize(self.bank_csv) == 0):
            print("[*] CSV files missing or empty. Generating fresh dataset...")
            from generate_datasets import main as gen_main
            gen_main()
            
        self.razorpay_df = pd.read_csv(self.razorpay_csv)
        self.bank_df = pd.read_csv(self.bank_csv)
        if os.path.exists(self.gst_csv):
            self.gst_df = pd.read_csv(self.gst_csv)
            
        # Standardize column names to lowercase stripped strings
        self.razorpay_df.columns = [str(c).lower().strip() for c in self.razorpay_df.columns]
        self.bank_df.columns = [str(c).lower().strip() for c in self.bank_df.columns]
        if not self.gst_df.empty:
            self.gst_df.columns = [str(c).lower().strip() for c in self.gst_df.columns]
            
        # Normalize alternative column aliases (e.g., rrn -> utr, txn_date -> date)
        if 'rrn' in self.razorpay_df.columns and 'utr' not in self.razorpay_df.columns:
            self.razorpay_df['utr'] = self.razorpay_df['rrn']
        if 'rrn' in self.bank_df.columns and 'utr' not in self.bank_df.columns:
            self.bank_df['utr'] = self.bank_df['rrn']
            
        # Ensure mandatory columns exist with default fallbacks
        for df, req_cols in [(self.razorpay_df, ['utr', 'payment_id', 'amount', 'date', 'merchant_id']),
                             (self.bank_df, ['utr', 'txn_ref', 'amount', 'value_date', 'merchant_id'])]:
            for col in req_cols:
                if col not in df.columns:
                    if col in ['utr']:
                        df[col] = ''
                    elif col in ['payment_id', 'txn_ref']:
                        df[col] = [f"id_{i}" for i in range(len(df))]
                    elif col in ['amount']:
                        df[col] = 0.0
                    elif col in ['date', 'value_date']:
                        df[col] = pd.Timestamp.now().strftime('%Y-%m-%d')
                    elif col in ['merchant_id']:
                        df[col] = 'MERCH_DEFAULT'

        # Convert date columns to datetime
        self.razorpay_df['date'] = pd.to_datetime(self.razorpay_df['date'], errors='coerce').fillna(pd.Timestamp.now())
        self.bank_df['value_date'] = pd.to_datetime(self.bank_df['value_date'], errors='coerce').fillna(pd.Timestamp.now())
        if not self.gst_df.empty and 'date' in self.gst_df.columns:
            self.gst_df['date'] = pd.to_datetime(self.gst_df['date'], errors='coerce').fillna(pd.Timestamp.now())
            
        # Clean UTR strings
        self.razorpay_df['utr'] = self.razorpay_df['utr'].fillna('').astype(str).str.strip()
        self.bank_df['utr'] = self.bank_df['utr'].fillna('').astype(str).str.strip()

    def run_tier1_exact_utr(self):
        """Tier 1: Exact UTR/RRN Matching."""
        rp_utr = self.razorpay_df[self.razorpay_df['utr'] != ''].copy()
        bank_utr = self.bank_df[self.bank_df['utr'] != ''].copy()
        
        bank_utr_counts = bank_utr['utr'].value_counts()
        duplicate_utrs = set(bank_utr_counts[bank_utr_counts > 1].index)
        
        clean_bank_utr = bank_utr[~bank_utr['utr'].isin(duplicate_utrs)]
        
        merged = pd.merge(
            rp_utr,
            clean_bank_utr,
            on='utr',
            suffixes=('_rp', '_bank')
        )
        
        matched_rp_ids = set()
        matched_bank_refs = set()
        
        for _, row in merged.iterrows():
            if abs(row['amount_rp'] - row['amount_bank']) <= 0.05:
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.matched_pairs.append({
                    'match_type': 'Tier 1 (Exact UTR)',
                    'payment_id': row['payment_id'],
                    'utr': row['utr'],
                    'merchant_id': row['merchant_id_rp'],
                    'razorpay_amount': row['amount_rp'],
                    'bank_amount': row['amount_bank'],
                    'razorpay_date': row['date'].strftime('%Y-%m-%d') if hasattr(row['date'], 'strftime') else str(row['date'])[:10],
                    'bank_date': row['value_date'].strftime('%Y-%m-%d') if hasattr(row['value_date'], 'strftime') else str(row['value_date'])[:10],
                    'confidence': 0.98,
                    'rule_fired': 'Exact UTR & Amount Match',
                    'timestamp': timestamp
                })
                matched_rp_ids.add(row['payment_id'])
                matched_bank_refs.add(row['txn_ref'])
                
        return matched_rp_ids, matched_bank_refs, duplicate_utrs

    def run_tier2_fuzzy(self, matched_rp_ids, matched_bank_refs):
        """Tier 2: Secondary Fuzzy Matching (Amount ±2%, Date T+1 to T+3, Merchant Match)."""
        unmatched_rp = self.razorpay_df[~self.razorpay_df['payment_id'].isin(matched_rp_ids)].copy()
        unmatched_bank = self.bank_df[~self.bank_df['txn_ref'].isin(matched_bank_refs)].copy()
        
        new_matched_rp = set()
        new_matched_bank = set()
        
        for _, rp_row in unmatched_rp.iterrows():
            best_match = None
            best_score = 0
            
            for _, bank_row in unmatched_bank.iterrows():
                if bank_row['txn_ref'] in new_matched_bank:
                    continue
                    
                if rp_row['merchant_id'] != bank_row['merchant_id']:
                    continue
                    
                amt_diff_pct = abs(rp_row['amount'] - bank_row['amount']) / (rp_row['amount'] if rp_row['amount'] > 0 else 1.0)
                if amt_diff_pct > 0.025:
                    continue
                    
                date_diff = (bank_row['value_date'] - rp_row['date']).days
                if date_diff < 0 or date_diff > 3:
                    continue
                    
                score = round(1.0 - (amt_diff_pct * 5) - (date_diff * 0.03), 2)
                score = max(0.75, min(0.94, score))
                
                if score > best_score:
                    best_score = score
                    best_match = bank_row
                    
            if best_match is not None:
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.matched_pairs.append({
                    'match_type': 'Tier 2 (Fuzzy Match)',
                    'payment_id': rp_row['payment_id'],
                    'utr': rp_row['utr'] or best_match['utr'] or 'N/A',
                    'merchant_id': rp_row['merchant_id'],
                    'razorpay_amount': rp_row['amount'],
                    'bank_amount': best_match['amount'],
                    'razorpay_date': rp_row['date'].strftime('%Y-%m-%d'),
                    'bank_date': best_match['value_date'].strftime('%Y-%m-%d'),
                    'confidence': best_score,
                    'rule_fired': f'Fuzzy Match (Amt Var: {round(abs(rp_row["amount"] - best_match["amount"]), 2)}, Date Window: T+{(best_match["value_date"] - rp_row["date"]).days})',
                    'timestamp': timestamp
                })
                new_matched_rp.add(rp_row['payment_id'])
                new_matched_bank.add(best_match['txn_ref'])
                
        return matched_rp_ids.union(new_matched_rp), matched_bank_refs.union(new_matched_bank)

    def run_tier3_exceptions(self, all_matched_rp_ids, all_matched_bank_refs, duplicate_utrs):
        """Tier 3: Exception Identification & AI Classification."""
        unmatched_rp = self.razorpay_df[~self.razorpay_df['payment_id'].isin(all_matched_rp_ids)].copy()
        unmatched_bank = self.bank_df[~self.bank_df['txn_ref'].isin(all_matched_bank_refs)].copy()
        
        # 1. Ghost/Duplicate UTRs
        for utr in duplicate_utrs:
            dup_rows = self.bank_df[self.bank_df['utr'] == utr]
            for _, bank_row in dup_rows.iterrows():
                if bank_row['txn_ref'] not in all_matched_bank_refs:
                    raw_rec = {
                        'exception_type': 'Ghost Entry / Duplicate',
                        'payment_id': f"bank_{bank_row['txn_ref']}",
                        'utr': utr,
                        'amount': bank_row['amount'],
                        'date': bank_row['value_date'].strftime('%Y-%m-%d'),
                        'merchant_id': bank_row['merchant_id']
                    }
                    ai_res = explain_exception(raw_rec)
                    raw_rec.update(ai_res)
                    self.exceptions.append(raw_rec)
                    all_matched_bank_refs.add(bank_row['txn_ref'])
        
        # 2. Unmatched Razorpay records
        for _, rp_row in unmatched_rp.iterrows():
            bank_candidates = unmatched_bank[
                (unmatched_bank['merchant_id'] == rp_row['merchant_id']) &
                (abs(unmatched_bank['amount'] - rp_row['amount']) <= 5.0)
            ]
            
            if not bank_candidates.empty:
                bank_match = bank_candidates.iloc[0]
                raw_rec = {
                    'exception_type': 'Timing Mismatch',
                    'payment_id': rp_row['payment_id'],
                    'utr': rp_row['utr'],
                    'amount': rp_row['amount'],
                    'razorpay_amount': rp_row['amount'],
                    'bank_amount': bank_match['amount'],
                    'razorpay_date': rp_row['date'].strftime('%Y-%m-%d'),
                    'bank_date': bank_match['value_date'].strftime('%Y-%m-%d'),
                    'date': rp_row['date'].strftime('%Y-%m-%d'),
                    'merchant_id': rp_row['merchant_id']
                }
            else:
                amt_candidates = unmatched_bank[
                    (unmatched_bank['merchant_id'] == rp_row['merchant_id']) &
                    (abs((unmatched_bank['value_date'] - rp_row['date']).dt.days) <= 3)
                ]
                if not amt_candidates.empty:
                    bank_match = amt_candidates.iloc[0]
                    raw_rec = {
                        'exception_type': 'Amount Mismatch',
                        'payment_id': rp_row['payment_id'],
                        'utr': rp_row['utr'],
                        'amount': rp_row['amount'],
                        'razorpay_amount': rp_row['amount'],
                        'bank_amount': bank_match['amount'],
                        'razorpay_date': rp_row['date'].strftime('%Y-%m-%d'),
                        'bank_date': bank_match['value_date'].strftime('%Y-%m-%d'),
                        'date': rp_row['date'].strftime('%Y-%m-%d'),
                        'merchant_id': rp_row['merchant_id']
                    }
                else:
                    raw_rec = {
                        'exception_type': 'Missing Entry (Bank)',
                        'payment_id': rp_row['payment_id'],
                        'utr': rp_row['utr'],
                        'amount': rp_row['amount'],
                        'date': rp_row['date'].strftime('%Y-%m-%d'),
                        'merchant_id': rp_row['merchant_id']
                    }
                    
            ai_res = explain_exception(raw_rec)
            raw_rec.update(ai_res)
            self.exceptions.append(raw_rec)

        # 3. Unmatched Bank records
        for _, bank_row in unmatched_bank.iterrows():
            if bank_row['txn_ref'] in all_matched_bank_refs:
                continue
            raw_rec = {
                'exception_type': 'Missing Entry (Razorpay)',
                'payment_id': f"unlinked_{bank_row['txn_ref']}",
                'txn_ref': bank_row['txn_ref'],
                'utr': bank_row['utr'],
                'amount': bank_row['amount'],
                'date': bank_row['value_date'].strftime('%Y-%m-%d'),
                'merchant_id': bank_row['merchant_id']
            }
            ai_res = explain_exception(raw_rec)
            raw_rec.update(ai_res)
            self.exceptions.append(raw_rec)

        # 4. GST Variances
        if not self.gst_df.empty:
            for _, gst_row in self.gst_df.iterrows():
                expected_gst = round(gst_row.get('taxable_amount', 0) * 0.18, 2)
                gst_amt = gst_row.get('gst_amount', 0)
                if abs(gst_amt - expected_gst) > 1.0:
                    raw_rec = {
                        'exception_type': 'GST Variance',
                        'payment_id': gst_row.get('payment_id', 'INV_VAR'),
                        'amount': gst_row.get('total_amount', 0),
                        'date': gst_row.get('date', pd.Timestamp.now()).strftime('%Y-%m-%d') if hasattr(gst_row.get('date'), 'strftime') else str(gst_row.get('date'))[:10],
                        'merchant_id': gst_row.get('merchant_id', 'MERCH_DEFAULT')
                    }
                    ai_res = explain_exception(raw_rec)
                    raw_rec.update(ai_res)
                    self.exceptions.append(raw_rec)

    def execute_reconciliation(self):
        """Run full 3-tier reconciliation workflow."""
        print("[*] Executing SettleIQ Reconciliation Engine...")
        self.load_data()
        
        matched_rp_ids, matched_bank_refs, dup_utrs = self.run_tier1_exact_utr()
        t1_count = len(self.matched_pairs)
        print(f"  [+] Tier 1 (Exact UTR): Matched {t1_count} records")
        
        all_matched_rp, all_matched_bank = self.run_tier2_fuzzy(matched_rp_ids, matched_bank_refs)
        t2_count = len(self.matched_pairs) - t1_count
        print(f"  [+] Tier 2 (Fuzzy Match): Matched {t2_count} additional records")
        
        self.run_tier3_exceptions(all_matched_rp, all_matched_bank, dup_utrs)
        print(f"  [+] Tier 3 (AI Exceptions): Classified {len(self.exceptions)} exceptions")
        
        matched_df = pd.DataFrame(self.matched_pairs)
        matched_df.to_csv('matched_pairs.csv', index=False)
        
        exceptions_df = pd.DataFrame(self.exceptions)
        exceptions_df.to_csv('exceptions.csv', index=False)
        
        self.export_excel_report(matched_df, exceptions_df)
        
        total_rp = len(self.razorpay_df)
        match_rate = round((len(matched_df) / total_rp) * 100, 1) if total_rp > 0 else 0
        
        summary = {
            'total_razorpay_records': total_rp,
            'total_matched': len(matched_df),
            'tier1_matches': t1_count,
            'tier2_matches': t2_count,
            'match_rate_pct': match_rate,
            'exceptions_count': len(exceptions_df),
            'total_amount_cleared': matched_df['bank_amount'].sum() if not matched_df.empty else 0,
            'amount_at_risk': exceptions_df['amount'].sum() if not exceptions_df.empty else 0
        }
        
        print("\n--- RECONCILIATION METRICS SUMMARY ---")
        print(f"Match Rate: {match_rate}% | Matched: {len(matched_df)} | Exceptions: {len(exceptions_df)}")
        print(f"Cleared Position: RS {summary['total_amount_cleared']:,.2f} | At Risk: RS {summary['amount_at_risk']:,.2f}")
        return summary

    def export_excel_report(self, matched_df, exceptions_df):
        """Export styled Excel file with green (matched) and red (exceptions) color coding."""
        try:
            excel_file = 'reconciliation_report.xlsx'
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                matched_df.to_excel(writer, sheet_name='Matched Records', index=False)
                exceptions_df.to_excel(writer, sheet_name='Exceptions Queue', index=False)
                
            import openpyxl
            from openpyxl.styles import PatternFill
            
            wb = openpyxl.load_workbook(excel_file)
            
            ws_matched = wb['Matched Records']
            green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            for row in ws_matched.iter_rows(min_row=2, max_col=ws_matched.max_column):
                for cell in row:
                    cell.fill = green_fill
                    
            ws_exc = wb['Exceptions Queue']
            red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            for row in ws_exc.iter_rows(min_row=2, max_col=ws_exc.max_column):
                for cell in row:
                    cell.fill = red_fill
                    
            wb.save(excel_file)
            print("[+] Exported color-coded Excel report -> reconciliation_report.xlsx")
        except Exception as e:
            print(f"[*] Excel formatting warning: {e}")

if __name__ == "__main__":
    engine = ReconciliationEngine()
    engine.execute_reconciliation()
